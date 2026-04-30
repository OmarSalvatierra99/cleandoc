"""
CleanDoc - Organizador de cédulas de resultados PDF
===================================================
Renombra y organiza PDFs por ente, periodo y anexos.
"""

import hashlib
import json
import logging
import re
import sqlite3
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Any, Dict

from werkzeug.datastructures import FileStorage

try:
    from utils import (
        FileProcessingError,
        InvalidFileError,
        sanitize_filename,
        validate_file_extension,
        validate_file_size,
    )
except ImportError:
    from .utils import (
        FileProcessingError,
        InvalidFileError,
        sanitize_filename,
        validate_file_extension,
        validate_file_size,
    )

try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None
    Font = None


logger = logging.getLogger(__name__)


PROJECTS_DIR = Path(__file__).resolve().parents[2]
CATALOGO_DB_CANDIDATES = (
    PROJECTS_DIR / "05-sasp" / "scil.db",
)
CATALOGO_GENERAL_CANDIDATES = (
    PROJECTS_DIR / "08-siif" / "catalogos" / "catalogo_general.json",
)

ANEXO_TYPES = {
    "PRAS": {
        "label": "PROMOCION DE RESPONSABILIDAD ADMINISTRATIVA SANCIONATORIA",
        "short_title": "PRAS_RESP_ADM",
        "patterns": (
            r"\(PRAS\)",
            r"PROMOCI(?:[ÓO]N|ONES)\s+DE RESPONSABILIDAD ADMINISTRATIVA SANCIONATORIA",
        ),
    },
    "PEFCF": {
        "label": "PROMOCION DEL EJERCICIO DE LA FACULTAD DE COMPROBACION FISCAL",
        "short_title": "PEFCF_COMP_FISCAL",
        "patterns": (
            r"\(PEFCF\)",
            r"PROMOCI(?:[ÓO]N|ONES)\s+DEL EJERCICIO DE LA FACULTAD DE COMPROBACI[ÓO]N FISCAL",
        ),
    },
    "PDP": {
        "label": "POSIBLE DANO PATRIMONIAL",
        "short_title": "PDP_DANO_PATR",
        "patterns": (
            r"\(PDP\)",
            r"POSIBLE DA[ÑN]O PATRIMONIAL",
            r"PROBABLE DA[ÑN]O PATRIMONIAL",
        ),
    },
    "SA": {
        "label": "SOLICITUD DE ACLARACION",
        "short_title": "SA_SOL_ACL",
        "patterns": (
            r"\(SA\)",
            r"SOLICITUD(?:ES)? DE ACLARACI[ÓO]N",
        ),
    },
    "R": {
        "label": "RECOMENDACION",
        "short_title": "R_RECOM",
        "patterns": (
            r"\(R\)",
            r"RECOMENDACI(?:[ÓO]N|ONES)",
        ),
    },
    "SIN_ANEXO": {
        "label": "SIN ANEXO",
        "short_title": "SIN_ANEXO",
        "patterns": (),
    },
}

ENTE_SIGLA_ALIAS_MAP = {
    "SM": "SMYT",
    "UPTREP": "UTREP",
    "UPT_REP": "UTREP",
}
CATALOGO_MATCH_THRESHOLD = 0.72


FUENTE_CODE_PATTERNS = (
    ("SREA", (r"SEGUIMIENTO A REMANENTES DE EJERCICIOS ANTERIORES",)),
    ("SPEA", (r"SEGUIMIENTO A PASIVOS DE EJERCICIOS ANTERIORES",)),
    ("REA", (r"REMANENTE(?:S)? DE EJERCICIOS ANTERIORES",)),
    ("IP", (r"INGRESOS PROPIOS",)),
    ("RFP", (r"RECURSOS FISCALES PROPIOS",)),
    ("RFE", (r"RECURSOS FISCALES ESTATALES",)),
    ("PE", (r"PARTICIPACIONES ESTATALES",)),
    ("RF", (r"RECURSOS FEDERALES",)),
    ("FGP", (r"FONDO GENERAL DE PARTICIPACIONES",)),
    ("ISR", (r"FONDO ISR",)),
    ("FF", (r"FONDO DE FISCALIZACI[ÓO]N",)),
    ("IDFL", (r"INGRESOS DERIVADOS DE FUENTES LOCALES",)),
    ("FAFEF", (r"FONDO DE APORTACIONES PARA EL FORTALECIMIENTO DE LAS ENTIDADES FEDERATIVAS",)),
    ("FAM", (r"FONDO DE APORTACIONES M[ÚU]LTIPLES",)),
    ("RR", (r"RECURSOS RECAUDADOS",)),
    ("CCTR", (r"CONVENIO DE COLABORACI[ÓO]N PARA LA TRANSFERENCIA DE RECURSOS",)),
    ("ST", (r"SALUD DE TLAXCALA",)),
    ("CONACYT", (r"CONACYT",)),
    ("CONADE", (r"CONADE",)),
    ("INEA", (r"INEA",)),
    ("EMSAD", (r"EMSAD",)),
)


FUENTE_CODE_OVERRIDES_RAW = {
    "Ingresos Propios": "IP",
    "Participaciones Estatales": "PE",
    "Participaciones Estatales (Fondo General de Participaciones)": "PE_FGP",
    "Participaciones Estatales (Fondo ISR)": "PE_ISR",
    "Participaciones Estatales (Ingresos derivados de fuentes locales)": "PE_IDFL",
    "Recursos Federales": "RF",
    "Recursos Fiscales Estatales": "RFE",
    "Recursos Fiscales Propios": "RFP",
    "Recursos Recaudados": "RR",
    "Recursos Recaudados y Participaciones Estatales": "RR_PE",
    "REA: Recursos Recaudados y Participaciones Estatales": "REA_RR_PE",
    "Remanente de Ejercicios Anteriores": "REA",
    "Remanentes de ejercicios anteriores": "REA",
    "Remanentes de Ejercicios Anteriores: Recursos Recaudados": "REA_RR",
    "Remanentes de Ejercicios Anteriores: Recursos Recaudados y Participaciones Estatales": "REA_RR_PE",
    "Seguimiento a Pasivos de Ejercicios Anteriores: Recursos Recaudados y Participaciones Estatales": "SPEA_RR_PE",
    "Seguimiento a Pasivos de Ejercicios Anteriores: Participaciones Estatales (Fondo General de Participaciones) 2023": "SPEA_PE_FGP",
    "Seguimiento a Pasivos de Ejercicios Anteriores: Participaciones Estatales (Fondo ISR) 2024": "SPEA_PE_ISR",
    "Seguimiento a Pasivos de Ejercicios Anteriores: Fondo de Aportaciones para el Fortalecimiento de las Entidades Federativas 2024": "SPEA_FAFEF",
    "Seguimiento a Remanentes de Ejercicios Anteriores: Participaciones Estatales (Fondo General de Compensación) 2024": "SREA_PE_FGC",
}


PERIODO_CODE_OVERRIDES_RAW = {
    "01 de enero al 31 de diciembre": "ENE-DIC",
    "01 de enero al 30 de septiembre": "ENE-SEP",
    "01 de enero al 31 de agosto": "ENE-AGO",
    "01 de enero al 30 de junio": "ENE-JUN",
    "01 de enero al 15 de mayo": "ENE-MAY",
    "16 de mayo al 30 de junio": "MAY-JUN",
    "01 de julio al 30 de septiembre": "JUL-SEP",
    "01 de julio al 31 de diciembre": "JUL-DIC",
    "01 de octubre al 31 de diciembre": "OCT-DIC",
}


@dataclass
class OrganizedPdfFile:
    original_name: str
    output_name: str
    archive_path: str
    ente_numero: str
    ente_codigo: str
    periodo_codigo: str
    fuente_codigo: str
    ente: str
    periodo: str
    fuente: str
    anexo: str
    ente_original: str
    periodo_original: str
    fuente_original: str
    observacion_refs: list[str]
    total_observaciones: int
    page_count: int
    text_hash: str
    stream: BytesIO


@dataclass
class PdfOrganizationStats:
    total_files: int = 0
    processed_files: int = 0
    duplicates_removed: int = 0
    skipped_files: int = 0
    folders_created: int = 0
    total_pages: int = 0
    total_observations: int = 0
    duplicate_names: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    organized_items: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "total_files": self.total_files,
            "processed_files": self.processed_files,
            "duplicates_removed": self.duplicates_removed,
            "skipped_files": self.skipped_files,
            "folders_created": self.folders_created,
            "total_pages": self.total_pages,
            "total_observations": self.total_observations,
            "duplicate_names": list(self.duplicate_names),
            "errors": list(self.errors),
            "organized_items": list(self.organized_items),
        }


def validar_pdf(file: FileStorage, max_size: int = 50 * 1024 * 1024) -> str:
    """Valida nombre, extensión y tamaño de un PDF."""
    if not file or not file.filename:
        raise InvalidFileError("No se proporcionó un archivo PDF válido")

    safe_filename = sanitize_filename(file.filename)
    validate_file_extension(safe_filename, allowed_extensions={".pdf"})
    validate_file_size(file, max_size)

    return safe_filename


def es_contenido_pdf_valido(file_stream) -> bool:
    """Verifica que el archivo tenga cabecera PDF."""
    try:
        file_stream.seek(0)
        header = file_stream.read(5)
        file_stream.seek(0)
        return header.startswith(b"%PDF-")
    except Exception:
        return False


def _construir_indices_relevantes_paginas(
    total_pages: int,
    head_pages: int = 6,
    tail_pages: int = 3,
) -> list[int]:
    """Devuelve índices de páginas iniciales/finales para una lectura rápida."""
    if total_pages <= 0:
        return []

    indices = set(range(min(head_pages, total_pages)))
    start_tail = max(total_pages - tail_pages, 0)
    indices.update(range(start_tail, total_pages))
    return sorted(indices)


def extraer_texto_pdf(
    file_stream,
    *,
    full_scan: bool = False,
    head_pages: int = 6,
    tail_pages: int = 3,
) -> tuple[str, int]:
    """Extrae texto de páginas clave del PDF y, si se pide, del documento completo."""
    if PdfReader is None:
        raise FileProcessingError("PyPDF2 no está instalado en el entorno")

    try:
        file_stream.seek(0)
        reader = PdfReader(file_stream)
        total_pages = len(reader.pages)
        fragmentos: list[str] = []

        if full_scan:
            page_indices = range(total_pages)
        else:
            page_indices = _construir_indices_relevantes_paginas(
                total_pages,
                head_pages=head_pages,
                tail_pages=tail_pages,
            )

        for page_index in page_indices:
            page = reader.pages[page_index]
            page_text = page.extract_text()
            if page_text:
                fragmentos.append(page_text)

        file_stream.seek(0)
        return "\n".join(fragmentos), total_pages
    except Exception as error:
        logger.warning("Error leyendo PDF: %s", error)
        file_stream.seek(0)
        return "", 0


def hash_texto(texto: str) -> str:
    """Genera hash para detectar duplicados."""
    return hashlib.md5(texto.encode("utf-8")).hexdigest()


def hash_pdf_stream(file_stream) -> str:
    """Genera hash binario del PDF para detectar duplicados exactos sin extraer texto completo."""
    file_stream.seek(0)
    digest = hashlib.md5()

    while True:
        chunk = file_stream.read(1024 * 1024)
        if not chunk:
            break
        digest.update(chunk)

    file_stream.seek(0)
    return digest.hexdigest()


def normalizar_espacios(texto: str) -> str:
    """Convierte saltos de línea y espacios múltiples en un solo espacio."""
    texto = texto.replace("\x0c", " ")
    return re.sub(r"\s+", " ", texto or "").strip()


def quitar_acentos(texto: str) -> str:
    """Convierte texto unicode a una versión ASCII simple."""
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(char)
    )


def normalizar_texto_clave(texto: str) -> str:
    """Normaliza texto libre para comparaciones exactas."""
    limpio = quitar_acentos(normalizar_espacios(texto)).lower()
    return re.sub(r"\s+", " ", limpio).strip()


FUENTE_CODE_OVERRIDES = {
    normalizar_texto_clave(raw): code
    for raw, code in FUENTE_CODE_OVERRIDES_RAW.items()
}
PERIODO_CODE_OVERRIDES = {
    normalizar_texto_clave(raw): code
    for raw, code in PERIODO_CODE_OVERRIDES_RAW.items()
}
ANEXO_LOOKAHEAD_PATTERN = (
    r"SOLICITUD(?:ES)? DE ACLARACI[ÓO]N|"
    r"RECOMENDACIONES?|"
    r"P(?:OSIBLE|ROBABLE) DA[ÑN]O PATRIMONIAL|"
    r"PROMOCI[ÓO]N(?:ES)? DE RESPONSABILIDAD ADMINISTRATIVA SANCIONATORIA|"
    r"PROMOCI[ÓO]N(?:ES)? DEL EJERCICIO DE LA FACULTAD DE COMPROBACI[ÓO]N FISCAL"
)


def limpiar_nombre(texto: str, max_len: int = 80) -> str:
    """Limpia texto para usarlo como nombre de archivo o carpeta."""
    texto = quitar_acentos(texto).upper()
    texto = re.sub(r"[^\w\s-]", "", texto)
    texto = re.sub(r"\s+", "_", texto.strip())
    texto = texto.strip("_")

    if not texto:
        return "SIN_DATO"

    return texto[:max_len].strip("_")


def normalizar_catalogo_clave(texto: str) -> str:
    """Normaliza nombres y siglas para resolver claves oficiales del catálogo."""
    limpio = quitar_acentos(normalizar_espacios(texto)).upper()
    limpio = re.sub(r"^\d+(?:\.\d+)*\s*[\.\)\-:]?\s*", "", limpio)
    limpio = re.sub(r"[^\w\s]", " ", limpio)
    return re.sub(r"\s+", " ", limpio).strip()


def normalizar_num_catalogo(texto: str) -> str:
    """Normaliza la clave numérica oficial del ente."""
    limpio = normalizar_espacios(texto or "")
    return re.sub(r"\.+$", "", limpio)


def _cargar_catalogo_desde_sasp_db() -> list[dict[str, str]]:
    """Lee entes y municipios desde la base institucional de SASP."""
    for db_path in CATALOGO_DB_CANDIDATES:
        if not db_path.exists():
            continue

        conn = None
        try:
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("""
                SELECT 'entes' AS grupo, num, nombre, siglas
                FROM entes
                WHERE activo=1
                UNION ALL
                SELECT 'municipios' AS grupo, num, nombre, siglas
                FROM municipios
                WHERE activo=1
            """)
            items = [
                {
                    "grupo": str(row["grupo"] or "").strip(),
                    "num": str(row["num"] or "").strip(),
                    "nombre": str(row["nombre"] or "").strip(),
                    "siglas": str(row["siglas"] or "").strip(),
                }
                for row in cur.fetchall()
            ]
            if items:
                return items
        except Exception as error:
            logger.warning(
                "No se pudo leer el catálogo desde SASP '%s': %s",
                db_path,
                error,
            )
        finally:
            if conn is not None:
                conn.close()

    return []


def _iter_catalogo_general_items() -> list[dict[str, str]]:
    """Carga el catálogo general más confiable disponible en proyectos vecinos."""
    items = _cargar_catalogo_desde_sasp_db()
    if items:
        return items

    for catalog_path in CATALOGO_GENERAL_CANDIDATES:
        if not catalog_path.exists():
            continue

        try:
            payload = json.loads(catalog_path.read_text(encoding="utf-8"))
        except Exception as error:
            logger.warning("No se pudo leer el catálogo general '%s': %s", catalog_path, error)
            continue

        items: list[dict[str, str]] = []
        for group_key in ("entes", "municipios"):
            for raw_item in payload.get(group_key, []):
                items.append(
                    {
                        "grupo": group_key,
                        "num": str(raw_item.get("num") or "").strip(),
                        "nombre": str(raw_item.get("nombre") or "").strip(),
                        "siglas": str(raw_item.get("siglas") or "").strip(),
                    }
                )
        if items:
            return items

    return []


@lru_cache(maxsize=1)
def construir_diccionario_entes() -> dict[str, dict[str, str]]:
    """Genera un índice normalizado hacia la ficha oficial del ente."""
    catalogo: dict[str, dict[str, str]] = {}
    por_sigla: dict[str, dict[str, str]] = {}

    def registrar(alias: str, record: dict[str, str]) -> None:
        alias_key = normalizar_catalogo_clave(alias)
        if alias_key:
            catalogo[alias_key] = record

    for item in _iter_catalogo_general_items():
        nombre = item["nombre"]
        siglas = limpiar_nombre(item["siglas"], max_len=24)
        siglas = ENTE_SIGLA_ALIAS_MAP.get(siglas, siglas)
        numero = normalizar_num_catalogo(item["num"])

        if not nombre or not siglas:
            continue

        record = {
            "num": numero,
            "siglas": siglas,
            "nombre": nombre,
        }
        por_sigla[siglas] = record

        for alias in (
            siglas,
            numero,
            f"{numero} {siglas}",
            f"{numero} {nombre}",
            nombre,
        ):
            registrar(alias, record)

    for alias, canonical in ENTE_SIGLA_ALIAS_MAP.items():
        record = por_sigla.get(canonical)
        if record:
            registrar(alias, record)

    return catalogo


@lru_cache(maxsize=1)
def construir_registros_entes() -> list[dict[str, str]]:
    """Devuelve registros únicos del catálogo para búsqueda aproximada."""
    seen: set[tuple[str, str, str]] = set()
    records: list[dict[str, str]] = []

    for record in construir_diccionario_entes().values():
        key = (
            record.get("num", ""),
            record.get("siglas", ""),
            record.get("nombre", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        records.append(record)

    return records


def construir_diccionario_anexos() -> dict[str, str]:
    """Expone el catálogo corto de tipos de anexo."""
    return {
        code: data["label"]
        for code, data in ANEXO_TYPES.items()
    }


def resolver_ente_catalogo(ente_original: str) -> dict[str, str] | None:
    """Busca la ficha oficial del ente en el catálogo general."""
    ente_key = normalizar_catalogo_clave(ente_original)
    if not ente_key:
        return None

    catalogo = construir_diccionario_entes()
    if ente_key in catalogo:
        return catalogo[ente_key]

    best_record = None
    best_score = 0.0

    for record in construir_registros_entes():
        candidates = (
            record.get("nombre", ""),
            record.get("siglas", ""),
            record.get("num", ""),
            f"{record.get('num', '')} {record.get('nombre', '')}",
            f"{record.get('num', '')} {record.get('siglas', '')}",
        )
        score = max(
            SequenceMatcher(
                None,
                ente_key,
                normalizar_catalogo_clave(candidate),
            ).ratio()
            for candidate in candidates
            if candidate
        )
        if score > best_score:
            best_score = score
            best_record = record

    if best_score >= CATALOGO_MATCH_THRESHOLD:
        return best_record

    return None


def describir_anexo(anexo: str) -> str:
    """Devuelve la etiqueta corta y legible del anexo."""
    return ANEXO_TYPES.get(anexo, ANEXO_TYPES["SIN_ANEXO"])["label"]


def construir_titulo_anexo(anexo: str) -> str:
    """Devuelve un título corto para nombrar archivos de anexo."""
    return ANEXO_TYPES.get(anexo, ANEXO_TYPES["SIN_ANEXO"])["short_title"]


def tokenizar_significativos(texto: str) -> list[str]:
    """Obtiene tokens útiles para construir claves cortas."""
    stopwords = {
        "DE", "DEL", "LA", "LAS", "EL", "LOS", "Y", "A", "AL", "PARA",
        "EN", "CON", "POR", "SE", "UN", "UNA", "SU", "SUS", "ETAPA",
    }

    limpio = quitar_acentos(texto).upper()
    tokens = re.findall(r"[A-Z0-9]+", limpio)
    return [token for token in tokens if token not in stopwords]


def construir_codigo_desde_tokens(tokens: list[str], default: str, max_tokens: int = 6) -> str:
    """Construye una clave corta a partir de tokens significativos."""
    if not tokens:
        return default

    acronym = "".join(token[0] for token in tokens[:max_tokens] if token)
    if 2 <= len(acronym) <= 12:
        return acronym

    combined = "_".join(tokens[:max_tokens])
    return combined[:24].strip("_") or default


def construir_codigo_ente(ente_original: str) -> str:
    """Construye una clave corta del ente para el nombre de archivo."""
    record = resolver_ente_catalogo(ente_original)
    if record and record.get("siglas"):
        return record["siglas"]

    tokens = tokenizar_significativos(ente_original)
    return construir_codigo_desde_tokens(tokens, "ENTE", max_tokens=8)


def construir_numero_ente(ente_original: str) -> str:
    """Obtiene el número institucional del ente."""
    record = resolver_ente_catalogo(ente_original)
    if record and record.get("num"):
        return record["num"]
    return "SIN_NUM"


def construir_codigo_periodo(periodo_original: str) -> str:
    """Construye una clave corta del periodo usando solo meses."""
    periodo_key = normalizar_texto_clave(periodo_original)
    if periodo_key in PERIODO_CODE_OVERRIDES:
        return PERIODO_CODE_OVERRIDES[periodo_key]

    meses = {
        "ENERO": "ENE",
        "FEBRERO": "FEB",
        "MARZO": "MAR",
        "ABRIL": "ABR",
        "MAYO": "MAY",
        "JUNIO": "JUN",
        "JULIO": "JUL",
        "AGOSTO": "AGO",
        "SEPTIEMBRE": "SEP",
        "SETIEMBRE": "SEP",
        "OCTUBRE": "OCT",
        "NOVIEMBRE": "NOV",
        "DICIEMBRE": "DIC",
    }

    limpio = quitar_acentos(periodo_original).upper()
    meses_detectados = [
        meses[token]
        for token in re.findall(
            r"ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|SETIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE",
            limpio,
        )
        if token in meses
    ]

    if meses_detectados:
        inicio = meses_detectados[0]
        fin = meses_detectados[-1]
        if inicio == fin:
            return inicio
        return f"{inicio}-{fin}"

    return limpiar_nombre(periodo_original, max_len=20).replace("_", "-")


def construir_codigo_fuente(fuente_original: str) -> str:
    """Construye una clave corta y práctica para la fuente de financiamiento."""
    fuente_key = normalizar_texto_clave(fuente_original)
    if fuente_key in FUENTE_CODE_OVERRIDES:
        return FUENTE_CODE_OVERRIDES[fuente_key]

    limpio = quitar_acentos(normalizar_espacios(fuente_original)).upper()
    matches: list[tuple[int, str]] = []

    for code, patterns in FUENTE_CODE_PATTERNS:
        positions = []
        for pattern in patterns:
            match = re.search(pattern, limpio, re.IGNORECASE)
            if match:
                positions.append(match.start())

        if positions:
            matches.append((min(positions), code))

    matches.sort()

    codigos: list[str] = []
    for _, code in matches:
        if code not in codigos:
            codigos.append(code)

    if codigos:
        return "_".join(codigos)

    tokens = tokenizar_significativos(fuente_original)
    return construir_codigo_desde_tokens(tokens, "FF", max_tokens=6)


def buscar_patron(texto: str, patron: str, default: str) -> str:
    """Extrae el primer valor que coincida con un patrón."""
    match = re.search(patron, texto, re.IGNORECASE)
    if not match:
        return default

    return normalizar_espacios(match.group(1))


def extraer_ente(texto: str) -> str:
    """Extrae el ente fiscalizable."""
    return buscar_patron(
        texto,
        rf"ENTE FISCALIZABLE:\s*(.+?)(?=(?:CONVENIO DE COLABORACI[ÓO]N|FUENTE DE FINANCIAMIENTO:|{ANEXO_LOOKAHEAD_PATTERN}|NO\.|REFERENCIA|P[ÓO]LIZA|FECHA|$))",
        "SIN_ENTE",
    )


def extraer_fuente(texto: str) -> str:
    """Extrae la fuente de financiamiento."""
    return buscar_patron(
        texto,
        rf"FUENTE DE FINANCIAMIENTO:\s*(.+?)(?=(?:NO\.|REFERENCIA|P[ÓO]LIZA|FECHA|{ANEXO_LOOKAHEAD_PATTERN}|PER[IÍ]ODO REVISADO:|MONTO OBSERVADO|DESCRIPCI[ÓO]N DEL RESULTADO|$))",
        "SIN_FUENTE",
    )


def extraer_periodo(texto: str) -> str:
    """Extrae el periodo revisado."""
    return buscar_patron(
        texto,
        r"PER[IÍ]ODO REVISADO:\s*(.+?)(?=(?:DESCRIPCI[ÓO]N DEL RESULTADO|MONTO OBSERVADO|NORMATIVIDAD|FECHA|CONCEPTO|BANCOS|NO\.|REFERENCIA|$))",
        "SIN_PERIODO",
    )


def extraer_anexo(texto: str) -> str:
    """Detecta el anexo a partir del nombre completo o su sigla."""
    for code, metadata in ANEXO_TYPES.items():
        if code == "SIN_ANEXO":
            continue

        patterns = metadata["patterns"]
        for pattern in patterns:
            if re.search(pattern, texto, re.IGNORECASE):
                return code

    return "SIN_ANEXO"


def extraer_observacion_refs(texto: str) -> list[str]:
    """Extrae referencias explícitas de observación."""
    refs = re.findall(r"OBSERVACI[ÓO]N:\s*([A-Z0-9.-]+)", texto, re.IGNORECASE)
    normalized = [normalizar_espacios(ref).upper() for ref in refs if ref.strip()]
    return sorted(set(normalized))


def extraer_total_observaciones(texto: str) -> int:
    """Extrae el total de observaciones reportado en el documento."""
    if not texto:
        return 0

    texto_con_lineas = (texto or "").replace("\r\n", "\n").replace("\r", "\n").replace("\x0c", "\n")
    matches = re.findall(
        r"^\s*(\d+)[ \t]+TOTAL\s+DE\s+OBSERVACIONES[.:]?\s*$",
        texto_con_lineas,
        re.IGNORECASE | re.MULTILINE,
    )
    if matches:
        return int(matches[-1])

    texto_normalizado = normalizar_espacios(texto_con_lineas)
    matches = re.findall(
        r"(\d+)\s+TOTAL\s+DE\s+OBSERVACIONES[.:]?",
        texto_normalizado,
        re.IGNORECASE,
    )
    if matches:
        return int(matches[-1])

    return 0


def extraer_campos(texto: str, filename: str = "") -> dict[str, Any]:
    """Extrae ente, fuente, periodo y anexo desde el texto normalizado."""
    texto_original = texto or ""
    texto_normalizado = normalizar_espacios(texto_original)

    ente_original = extraer_ente(texto_normalizado)
    ente_match = resolver_ente_catalogo(ente_original)
    if not ente_match and filename:
        ente_match = resolver_ente_catalogo(Path(filename).stem)
    if ente_match:
        ente_original = ente_match["nombre"]
    fuente_original = extraer_fuente(texto_normalizado)
    periodo_original = extraer_periodo(texto_normalizado)
    anexo = extraer_anexo(texto_normalizado)
    observacion_refs = extraer_observacion_refs(texto_normalizado)
    total_observaciones = extraer_total_observaciones(texto_original)

    return {
        "ente_original": ente_original,
        "fuente_original": fuente_original,
        "periodo_original": periodo_original,
        "anexo": anexo,
        "ente_resuelto": bool(ente_match),
        "observacion_refs": observacion_refs,
        "total_observaciones": total_observaciones,
        "ente": limpiar_nombre(ente_original, max_len=70),
        "periodo": limpiar_nombre(periodo_original, max_len=40),
        "fuente": limpiar_nombre(fuente_original, max_len=70),
        "ente_numero": (ente_match or {}).get("num") or construir_numero_ente(ente_original),
        "ente_codigo": (ente_match or {}).get("siglas") or construir_codigo_ente(ente_original),
        "periodo_codigo": construir_codigo_periodo(periodo_original),
        "fuente_codigo": construir_codigo_fuente(fuente_original),
    }


def describir_incidencia_campos(campos: dict[str, Any]) -> str:
    """Genera una incidencia legible cuando faltan datos clave para renombrar."""
    incidencias: list[str] = []

    ente_detectado = normalizar_espacios(campos.get("ente_original", ""))
    if ente_detectado == "SIN_ENTE":
        incidencias.append("no se detectó el campo 'ENTE FISCALIZABLE'")
    elif not campos.get("ente_resuelto") or not campos.get("ente_numero"):
        incidencias.append(
            f"el ente detectado '{ente_detectado}' no se pudo vincular con el catálogo oficial"
        )

    fuente_detectada = normalizar_espacios(campos.get("fuente_original", ""))
    if fuente_detectada == "SIN_FUENTE":
        incidencias.append("no se detectó el campo 'FUENTE DE FINANCIAMIENTO'")

    periodo_detectado = normalizar_espacios(campos.get("periodo_original", ""))
    if periodo_detectado == "SIN_PERIODO":
        incidencias.append("no se detectó el campo 'PERÍODO REVISADO'")

    if campos.get("anexo") == "SIN_ANEXO":
        incidencias.append(
            "no se identificó el tipo de anexo (SA, R, PDP, PRAS o PEFCF)"
        )

    if incidencias:
        return "; ".join(incidencias)

    return "no se pudieron obtener datos válidos para construir el nombre de salida"


def _faltan_campos_clave(campos: dict[str, Any]) -> bool:
    """Indica si faltan datos mínimos para construir el nombre final."""
    return (
        not campos.get("ente_resuelto")
        or not campos.get("ente_numero")
        or campos.get("fuente_original") == "SIN_FUENTE"
        or campos.get("periodo_original") == "SIN_PERIODO"
        or campos.get("anexo") == "SIN_ANEXO"
    )


def construir_nombre_salida(
    ente_numero: str,
    ente_codigo: str,
    fuente_codigo: str,
    periodo_codigo: str,
    anexo: str,
) -> str:
    """Construye el nombre base del archivo renombrado."""
    return f"{ente_numero}_{ente_codigo}_{fuente_codigo}_{periodo_codigo}_{anexo}"


def construir_ruta_unica(
    ente_numero: str,
    ente_codigo: str,
    periodo_codigo: str,
    fuente_codigo: str,
    anexo: str,
    usados: set[str],
) -> tuple[str, str]:
    """Genera un nombre de salida único en la raíz del ZIP."""
    base_nombre = construir_nombre_salida(
        ente_numero,
        ente_codigo,
        fuente_codigo,
        periodo_codigo,
        anexo,
    )
    output_name = f"{base_nombre}.pdf"
    archive_path = output_name
    contador = 1

    while archive_path in usados:
        output_name = f"{base_nombre}_{contador}.pdf"
        archive_path = output_name
        contador += 1

    return output_name, archive_path


def agregar_hoja_resumen(workbook, organized_items: list[dict[str, Any]]) -> None:
    """Agrega hoja resumen agrupada por ente, periodo, fuente y anexo."""
    summary_sheet = workbook.create_sheet("Resumen")
    headers = [
        "ENTE_CODIGO",
        "PERIODO_CODIGO",
        "FF_CODIGO",
        "ENTE",
        "PERIODO",
        "FUENTE_DE_FINANCIAMIENTO",
        "ANEXO",
        "TOTAL_OBSERVACIONES_AL_FINAL",
        "TOTAL_PDFS",
        "TOTAL_PAGINAS",
    ]
    summary_sheet.append(headers)

    grouped = Counter(
        (
            item["ente_codigo"],
            item["periodo_codigo"],
            item["fuente_codigo"],
            item["ente_original"],
            item["periodo_original"],
            item["fuente_original"],
            item["anexo"],
        )
        for item in organized_items
    )

    pages_by_group = Counter()
    observations_by_group = Counter()

    for item in organized_items:
        key = (
            item["ente_codigo"],
            item["periodo_codigo"],
            item["fuente_codigo"],
            item["ente_original"],
            item["periodo_original"],
            item["fuente_original"],
            item["anexo"],
        )
        pages_by_group[key] += item["page_count"]
        observations_by_group[key] += item["total_observaciones"]

    for key in sorted(grouped):
        summary_sheet.append([
            key[0],
            key[1],
            key[2],
            key[3],
            key[4],
            key[5],
            key[6],
            observations_by_group[key],
            grouped[key],
            pages_by_group[key],
        ])


def agregar_hoja_detalle(workbook, organized_items: list[dict[str, Any]]) -> None:
    """Agrega hoja detalle con todos los archivos procesados."""
    detail_sheet = workbook.active
    detail_sheet.title = "Detalle"
    headers = [
        "ARCHIVO_ORIGINAL",
        "ENTE_CODIGO",
        "PERIODO_CODIGO",
        "FF_CODIGO",
        "ENTE",
        "PERIODO",
        "FUENTE_DE_FINANCIAMIENTO",
        "ANEXO",
        "OBSERVACIONES_REFERIDAS",
        "TOTAL_OBSERVACIONES_AL_FINAL",
        "PAGINAS",
        "ARCHIVO_SALIDA",
        "RUTA_EN_ZIP",
    ]
    detail_sheet.append(headers)

    for item in organized_items:
        detail_sheet.append([
            item["original_name"],
            item["ente_codigo"],
            item["periodo_codigo"],
            item["fuente_codigo"],
            item["ente_original"],
            item["periodo_original"],
            item["fuente_original"],
            item["anexo"],
            ", ".join(item["observacion_refs"]),
            item["total_observaciones"],
            item["page_count"],
            item["output_name"],
            item["archive_path"],
        ])


def agregar_hoja_duplicados(workbook, duplicate_names: list[str]) -> None:
    """Agrega hoja con duplicados omitidos."""
    duplicates_sheet = workbook.create_sheet("Duplicados")
    duplicates_sheet.append(["ARCHIVO_DUPLICADO"])

    for duplicate_name in duplicate_names:
        duplicates_sheet.append([duplicate_name])


def agregar_hoja_errores(workbook, errors: list[str]) -> None:
    """Agrega hoja con incidencias de procesamiento."""
    errors_sheet = workbook.create_sheet("Errores")
    errors_sheet.append(["INCIDENCIA"])

    for error in errors:
        errors_sheet.append([error])


def agregar_hoja_conteo_observaciones(workbook, organized_items: list[dict[str, Any]]) -> None:
    """Agrega hoja enfocada en el conteo final de observaciones."""
    observations_sheet = workbook.create_sheet("ConteoObservaciones")
    headers = [
        "ENTE_CODIGO",
        "PERIODO_CODIGO",
        "FF_CODIGO",
        "ANEXO",
        "TOTAL_OBSERVACIONES_AL_FINAL",
        "TOTAL_PDFS",
        "ENTE",
        "PERIODO",
        "FUENTE_DE_FINANCIAMIENTO",
    ]
    observations_sheet.append(headers)

    grouped = Counter(
        (
            item["ente_codigo"],
            item["periodo_codigo"],
            item["fuente_codigo"],
            item["anexo"],
            item["ente_original"],
            item["periodo_original"],
            item["fuente_original"],
        )
        for item in organized_items
    )
    observations_by_group = Counter()

    for item in organized_items:
        key = (
            item["ente_codigo"],
            item["periodo_codigo"],
            item["fuente_codigo"],
            item["anexo"],
            item["ente_original"],
            item["periodo_original"],
            item["fuente_original"],
        )
        observations_by_group[key] += item["total_observaciones"]

    for key in sorted(grouped):
        observations_sheet.append([
            key[0],
            key[1],
            key[2],
            key[3],
            observations_by_group[key],
            grouped[key],
            key[4],
            key[5],
            key[6],
        ])


def agregar_hoja_claves(workbook, organized_items: list[dict[str, Any]]) -> None:
    """Agrega catálogo de claves cortas para validar la organización."""
    codes_sheet = workbook.create_sheet("Claves")
    codes_sheet.append(["TIPO", "CODIGO", "VALOR_COMPLETO"])

    seen: set[tuple[str, str, str]] = set()

    for item in organized_items:
        rows = (
            ("ENTE", item["ente_codigo"], item["ente_original"]),
            ("PERIODO", item["periodo_codigo"], item["periodo_original"]),
            ("FF", item["fuente_codigo"], item["fuente_original"]),
            ("ANEXO", item["anexo"], describir_anexo(item["anexo"])),
        )

        for row in rows:
            if row in seen:
                continue
            seen.add(row)
            codes_sheet.append(list(row))


def dar_formato_libro(workbook) -> None:
    """Aplica formato básico al libro Excel."""
    if Font is None:
        return

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        sheet.freeze_panes = "A2"

        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def construir_excel_resumen(stats: PdfOrganizationStats) -> BytesIO:
    """Construye un resumen Excel para validar la organización."""
    if Workbook is None:
        raise FileProcessingError("openpyxl no está instalado en el entorno")

    workbook = Workbook()

    agregar_hoja_detalle(workbook, stats.organized_items)
    agregar_hoja_resumen(workbook, stats.organized_items)
    agregar_hoja_conteo_observaciones(workbook, stats.organized_items)
    agregar_hoja_claves(workbook, stats.organized_items)
    agregar_hoja_duplicados(workbook, stats.duplicate_names)
    agregar_hoja_errores(workbook, stats.errors)
    dar_formato_libro(workbook)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def organizar_cedulas_resultados_pdf(
    files: list[FileStorage],
    max_size: int = 50 * 1024 * 1024,
) -> tuple[list[OrganizedPdfFile], PdfOrganizationStats]:
    """Procesa PDFs, elimina duplicados y los organiza por ente, periodo y anexos."""
    stats = PdfOrganizationStats(total_files=len(files))
    hashes: dict[str, str] = {}
    rutas_usadas: set[str] = set()
    organized_files: list[OrganizedPdfFile] = []

    for file in files:
        if not file or not file.filename:
            stats.skipped_files += 1
            stats.errors.append("Se recibió un archivo vacío")
            continue

        try:
            safe_filename = validar_pdf(file, max_size)

            if not es_contenido_pdf_valido(file.stream):
                raise InvalidFileError("El archivo no es un PDF válido")

            texto, page_count = extraer_texto_pdf(file.stream)
            if not texto.strip():
                stats.skipped_files += 1
                stats.errors.append(
                    f"{safe_filename}: No se pudo extraer texto legible del PDF"
                )
                continue

            texto_hash = hash_pdf_stream(file.stream)

            if texto_hash in hashes:
                stats.duplicates_removed += 1
                stats.duplicate_names.append(safe_filename)
                continue

            hashes[texto_hash] = safe_filename

            campos = extraer_campos(texto, filename=safe_filename)
            if _faltan_campos_clave(campos):
                texto_completo, full_page_count = extraer_texto_pdf(file.stream, full_scan=True)
                if texto_completo.strip():
                    texto = texto_completo
                    page_count = full_page_count
                    campos = extraer_campos(texto, filename=safe_filename)

            if _faltan_campos_clave(campos):
                raise InvalidFileError(describir_incidencia_campos(campos))
            output_name, archive_path = construir_ruta_unica(
                campos["ente_numero"],
                campos["ente_codigo"],
                campos["periodo_codigo"],
                campos["fuente_codigo"],
                campos["anexo"],
                rutas_usadas,
            )

            rutas_usadas.add(archive_path)

            file.stream.seek(0)
            pdf_bytes = file.stream.read()
            output_stream = BytesIO(pdf_bytes)
            output_stream.seek(0)

            organized_files.append(
                OrganizedPdfFile(
                    original_name=safe_filename,
                    output_name=output_name,
                    archive_path=archive_path,
                    ente_numero=campos["ente_numero"],
                    ente_codigo=campos["ente_codigo"],
                    periodo_codigo=campos["periodo_codigo"],
                    fuente_codigo=campos["fuente_codigo"],
                    ente=campos["ente"],
                    periodo=campos["periodo"],
                    fuente=campos["fuente"],
                    anexo=campos["anexo"],
                    ente_original=campos["ente_original"],
                    periodo_original=campos["periodo_original"],
                    fuente_original=campos["fuente_original"],
                    observacion_refs=campos["observacion_refs"],
                    total_observaciones=campos["total_observaciones"],
                    page_count=page_count,
                    text_hash=texto_hash,
                    stream=output_stream,
                )
            )

            stats.processed_files += 1
            stats.total_pages += page_count
            stats.total_observations += campos["total_observaciones"]
            stats.organized_items.append(
                {
                    "original_name": safe_filename,
                    "output_name": output_name,
                    "archive_path": archive_path,
                    "ente_numero": campos["ente_numero"],
                    "ente_codigo": campos["ente_codigo"],
                    "periodo_codigo": campos["periodo_codigo"],
                    "fuente_codigo": campos["fuente_codigo"],
                    "ente": campos["ente"],
                    "periodo": campos["periodo"],
                    "fuente": campos["fuente"],
                    "anexo": campos["anexo"],
                    "anexo_titulo": describir_anexo(campos["anexo"]),
                    "ente_original": campos["ente_original"],
                    "periodo_original": campos["periodo_original"],
                    "fuente_original": campos["fuente_original"],
                    "observacion_refs": campos["observacion_refs"],
                    "total_observaciones": campos["total_observaciones"],
                    "page_count": page_count,
                }
            )
        except InvalidFileError as error:
            stats.skipped_files += 1
            stats.errors.append(f"{file.filename}: {error.message}")
        except Exception as error:
            stats.skipped_files += 1
            logger.error("Error organizando PDF '%s': %s", file.filename, error, exc_info=True)
            stats.errors.append(f"{file.filename}: {error}")

    stats.folders_created = 0

    return organized_files, stats
